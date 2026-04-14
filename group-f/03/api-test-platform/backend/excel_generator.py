from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import json


def generate_excel(project, interfaces, test_cases):
    """生成Excel测试用例文档"""
    wb = Workbook()
    
    # Sheet 1: 项目信息
    ws_info = wb.active
    ws_info.title = "项目信息"
    _create_project_info_sheet(ws_info, project)
    
    # Sheet 2: 接口清单
    ws_interfaces = wb.create_sheet("接口清单")
    _create_interfaces_sheet(ws_interfaces, interfaces)
    
    # Sheet 3: 测试用例详情
    ws_cases = wb.create_sheet("测试用例详情")
    _create_test_cases_sheet(ws_cases, test_cases)
    
    # 保存文件
    filename = f"testcases_{project.name.lower().replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = f"generated_tests/{filename}"
    wb.save(filepath)
    
    return filepath, filename


def _create_project_info_sheet(ws, project):
    """创建项目信息Sheet"""
    # 标题
    ws['A1'] = "项目信息"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:B1')
    
    # 数据
    data = [
        ("项目名称", project.name),
        ("项目描述", project.description or ""),
        ("基础URL", project.base_url),
        ("导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    
    # 表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for idx, (key, value) in enumerate(data, start=3):
        ws[f'A{idx}'] = key
        ws[f'B{idx}'] = value
        ws[f'A{idx}'].font = header_font
        ws[f'A{idx}'].fill = header_fill
    
    # 调整列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 50


def _create_interfaces_sheet(ws, interfaces):
    """创建接口清单Sheet"""
    # 表头
    headers = ["序号", "接口名称", "请求方法", "接口路径", "用例数量"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # 数据
    for idx, interface in enumerate(interfaces, start=2):
        ws.cell(row=idx, column=1, value=idx - 1)
        ws.cell(row=idx, column=2, value=interface.name)
        ws.cell(row=idx, column=3, value=interface.method)
        ws.cell(row=idx, column=4, value=interface.path)
        ws.cell(row=idx, column=5, value=len(interface.test_cases))
    
    # 调整列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 12


def _create_test_cases_sheet(ws, test_cases):
    """创建测试用例详情Sheet"""
    # 表头
    headers = ["序号", "所属接口", "用例名称", "用例描述", "请求方法", "请求URL", 
               "请求头", "请求参数", "请求体", "期望状态码", "期望响应"]
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 数据
    for idx, case in enumerate(test_cases, start=2):
        interface = case.interface
        project = interface.project
        
        ws.cell(row=idx, column=1, value=idx - 1)
        ws.cell(row=idx, column=2, value=interface.name)
        ws.cell(row=idx, column=3, value=case.name)
        ws.cell(row=idx, column=4, value=case.description or "")
        ws.cell(row=idx, column=5, value=interface.method)
        ws.cell(row=idx, column=6, value=f"{project.base_url}{interface.path}")
        ws.cell(row=idx, column=7, value=json.dumps(interface.headers or {}, ensure_ascii=False))
        ws.cell(row=idx, column=8, value=json.dumps(case.request_params or {}, ensure_ascii=False))
        ws.cell(row=idx, column=9, value=json.dumps(case.request_body or {}, ensure_ascii=False))
        ws.cell(row=idx, column=10, value=case.expected_status)
        ws.cell(row=idx, column=11, value=json.dumps(case.expected_response or {}, ensure_ascii=False))
    
    # 调整列宽
    col_widths = [8, 20, 20, 30, 12, 50, 30, 30, 30, 12, 30]
    for idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # 设置所有单元格样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)